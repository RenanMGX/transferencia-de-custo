import os
import pandas as pd
import openpyxl
from tkinter import filedialog
import json
import datetime
import getpass

class Config():
    def __init__(self):
        self.__caminho = f"C:\\Users\\{getpass.getuser()}\\.bot_transferencia_custo"
        self.__caminho_config = f"{self.__caminho}\\config.json"
        self.__config_temp = {
            "cadastro_de_empresas" : ""
        }

        if not os.path.exists(self.__caminho):
            os.makedirs(self.__caminho)
    
    def load(self):
        if os.path.exists(self.__caminho_config):
            with open(self.__caminho_config, 'r')as arqui:
                return json.load(arqui)
        else:
            with open(self.__caminho_config, 'w')as arqui:
                json.dump(self.__config_temp, arqui)

class Robo():
    def __init__(self):
        self.__config = configuracoes.load()
        self.__lista_de_arquivos = []
        self.dados_do_formulario_transferencia = []

        self.data_documento = datetime.datetime.now().strftime("%d.%m.%Y")
        self.data_vencimento = datetime.datetime.now().strftime("23.%m.%Y")


    def listar_arquivos(self):
        self.__pasta = filedialog.askdirectory()
        self.__lista_de_arquivos = list(os.listdir(self.__pasta))
        for indice,arquivo in enumerate(self.__lista_de_arquivos):
            if arquivo[0] == "~":
                self.__lista_de_arquivos.pop(indice)
            else:
                self.__lista_de_arquivos[indice] = f"{self.__pasta}/{arquivo}"
    
    def carregar_cadastro_de_empresas(self):
        caminho = self.__config['cadastro_de_empresas']
        self.cadastro_de_empresas = pd.read_excel(caminho, header=1)
        

    def carregar_arquivos_da_lista(self):
        for arquivo in self.__lista_de_arquivos:
            if (".xlsx" in arquivo.lower()) or (".xlsm" in arquivo.lower()) or (".xlsb" in arquivo.lower()) or (".xltx" in arquivo.lower()):
                dados = {}
                try: 
                    wb = openpyxl.load_workbook(arquivo, data_only=True)
                except PermissionError:
                    print(f"{arquivo} está aberto em outro programa")
                    continue

                ws = wb.active

                #verifica se é o tipo de planilha certa
                print(ws['B2'].value)
                if ws['B2'].value != "FORMULÁRIO DE TRANSFERÊNCIA DE CUSTOS":
                    continue

                dados['divisao_origem'] = ws['D8'].value
                dados['divisao_destino'] = ws['J8'].value

                dados["linhas"] = []
                lancamentos = ws['B17:K467']
                for row in lancamentos:
                    lista = {}
                    if row[0].value == None:
                        continue

                    lista['origem_tipo'] = row[0].value
                    lista['origem_conta_do_razao'] = row[1].value
                    lista['origem_debito_credito'] = row[2].value
                    lista['origem_pep_centro_de_custo_empresa_origem'] = row[3].value
                    lista['destino_tipo'] = row[4].value
                    lista['destino_conta_do_razao'] = row[5].value
                    lista['destino_debito_credito'] = row[6].value
                    lista['destino_pep_centro_de_custo_empresa_origem'] = row[7].value
                    lista['valor'] = row[8].value
                    lista['descricao'] = row[9].value

                    dados['linhas'].append(lista)

                self.dados_do_formulario_transferencia.append(dados)
                self.montar_dados()
    
    def montar_dados(self):
        linhas_temp = []
        sequencial = 1
        for dados_brutos in self.dados_do_formulario_transferencia:
            for dados_linha in dados_brutos['linhas']:
                linhas_montagem = []

                ############## Linha 1
                sequencial_demo = "0000" + str(sequencial)
                sequencial_demo = sequencial_demo[-4:]
                linhas_montagem.append(sequencial_demo) # sequencial
                linhas_montagem.append(self.data_documento)    #data do documento
                linhas_montagem.append(self.data_documento)    #data do documento
                linhas_montagem.append(self.cadastro_de_empresas[self.cadastro_de_empresas['Divisão'] == dados_brutos['divisao_origem']]['Empresa'].values[0])  # transforma a divisão d empresa na empresa
                linhas_montagem.append(dados_brutos['divisao_origem'])  #divisão da empresa
                linhas_montagem.append("AB") ############ tipo do documento #### olhar com a Rafaela
                linhas_montagem.append("Nota de Débito") ############ Texto cabeçalho #### olhar com a Rafaela
                linhas_montagem.append("") ############ Referencia #### olhar com a Rafaela
                linhas_montagem.append("") ############ Cód. Rze  #### olhar com a Rafaela
                linhas_montagem.append(50) ############ Chave de laçamento  #### olhar com a Rafaela
                linhas_montagem.append(dados_linha['valor']) #Valor
                linhas_montagem.append("s") ############ Tipo de Conta  #### olhar com a Rafaela
                linhas_montagem.append(int(dados_linha['origem_conta_do_razao'])) #Valor

                if "." in dados_linha['origem_pep_centro_de_custo_empresa_origem']:
                    linhas_montagem.append("") #Centro de Custo
                    linhas_montagem.append(dados_linha['origem_pep_centro_de_custo_empresa_origem']) #PEP
                else:
                    linhas_montagem.append(dados_linha['origem_pep_centro_de_custo_empresa_origem']) #Centro de Custo
                    linhas_montagem.append("") #PEP
                
                linhas_montagem.append("") ############ Ordem  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Centro de Lucro  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Tipo de Atividade  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Data Vencimento  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Atribuicao  #### olhar com a Rafaela
                linhas_montagem.append(dados_linha['descricao']) #Histórico
                
                ############## Saltando Linha
                linhas_temp.append(linhas_montagem)
                linhas_montagem = []

                ############## Linha 2
                sequencial_demo = "0000" + str(sequencial)
                sequencial_demo = sequencial_demo[-4:]
                linhas_montagem.append(sequencial_demo) # sequencial
                linhas_montagem.append("")    #data do documento
                linhas_montagem.append("")    #data do documento
                linhas_montagem.append("")  # transforma a divisão d empresa na empresa
                linhas_montagem.append(dados_brutos['divisao_origem'])  #divisão da empresa
                linhas_montagem.append("") ############ tipo do documento #### olhar com a Rafaela
                linhas_montagem.append("") ############ Texto cabeçalho #### olhar com a Rafaela
                linhas_montagem.append("") ############ Referencia #### olhar com a Rafaela
                linhas_montagem.append("") ############ Cód. Rze  #### olhar com a Rafaela
                linhas_montagem.append(40) ############ Chave de laçamento  #### olhar com a Rafaela
                linhas_montagem.append(dados_linha['valor']) #Valor
                linhas_montagem.append("s") ############ Tipo de Conta  #### olhar com a Rafaela
                linhas_montagem.append(int(self.cadastro_de_empresas[self.cadastro_de_empresas['Divisão'] == dados_brutos['divisao_destino']]['Conta '].values[0])) #Valor

                linhas_montagem.append("") #Centro de Custo
                linhas_montagem.append("") #PEP
                
                linhas_montagem.append("") ############ Ordem  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Centro de Lucro  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Tipo de Atividade  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Data Vencimento  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Atribuicao  #### olhar com a Rafaela
                linhas_montagem.append(f"ND {dados_linha['descricao']}") #Histórico

                ############## Saltando Linha
                linhas_temp.append(linhas_montagem)
                sequencial += 1
                linhas_montagem = []

                ############## Linha 3
                sequencial_demo = "0000" + str(sequencial)
                sequencial_demo = sequencial_demo[-4:]
                linhas_montagem.append(sequencial_demo) # sequencial
                linhas_montagem.append(self.data_documento)    #data do documento
                linhas_montagem.append(self.data_documento)    #data do documento
                linhas_montagem.append(self.cadastro_de_empresas[self.cadastro_de_empresas['Divisão'] == dados_brutos['divisao_destino']]['Empresa'].values[0])  # transforma a divisão d empresa na empresa
                linhas_montagem.append(dados_brutos['divisao_destino'])  #divisão da empresa
                linhas_montagem.append("AB") ############ tipo do documento #### olhar com a Rafaela
                linhas_montagem.append("Nota de Débito") ############ Texto cabeçalho #### olhar com a Rafaela
                linhas_montagem.append("") ############ Referencia #### olhar com a Rafaela
                linhas_montagem.append("") ############ Cód. Rze  #### olhar com a Rafaela
                linhas_montagem.append(40) ############ Chave de laçamento  #### olhar com a Rafaela
                linhas_montagem.append(dados_linha['valor']) #Valor
                linhas_montagem.append("s") ############ Tipo de Conta  #### olhar com a Rafaela
                linhas_montagem.append(int(dados_linha['destino_conta_do_razao'])) #Valor

                if "." in dados_linha['destino_pep_centro_de_custo_empresa_origem']:
                    linhas_montagem.append("") #Centro de Custo
                    linhas_montagem.append(dados_linha['destino_pep_centro_de_custo_empresa_origem']) #PEP
                else:
                    linhas_montagem.append(dados_linha['destino_pep_centro_de_custo_empresa_origem']) #Centro de Custo
                    linhas_montagem.append("") #PEP
                
                linhas_montagem.append("") ############ Ordem  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Centro de Lucro  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Tipo de Atividade  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Data Vencimento  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Atribuicao  #### olhar com a Rafaela
                linhas_montagem.append(dados_linha['descricao']) #Histórico

                ############## Saltando Linha
                linhas_temp.append(linhas_montagem)
                linhas_montagem = []

                ############## Linha 4
                sequencial_demo = "0000" + str(sequencial)
                sequencial_demo = sequencial_demo[-4:]
                linhas_montagem.append(sequencial_demo) # sequencial
                linhas_montagem.append("")    #data do documento
                linhas_montagem.append("")    #data do documento
                linhas_montagem.append("")  # transforma a divisão d empresa na empresa
                linhas_montagem.append(dados_brutos['divisao_destino'])  #divisão da empresa
                linhas_montagem.append("") ############ tipo do documento #### olhar com a Rafaela
                linhas_montagem.append("") ############ Texto cabeçalho #### olhar com a Rafaela
                linhas_montagem.append("") ############ Referencia #### olhar com a Rafaela
                linhas_montagem.append("") ############ Cód. Rze  #### olhar com a Rafaela
                linhas_montagem.append(40) ############ Chave de laçamento  #### olhar com a Rafaela
                linhas_montagem.append(dados_linha['valor']) #Valor
                linhas_montagem.append("s") ############ Tipo de Conta  #### olhar com a Rafaela
                linhas_montagem.append(int(self.cadastro_de_empresas[self.cadastro_de_empresas['Divisão'] == dados_brutos['divisao_origem']]['Código '].values[0])) #Valor

                linhas_montagem.append("") #Centro de Custo
                linhas_montagem.append("") #PEP
                
                linhas_montagem.append("") ############ Ordem  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Centro de Lucro  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Tipo de Atividade  #### olhar com a Rafaela
                linhas_montagem.append(self.data_vencimento) ############ Data Vencimento  #### olhar com a Rafaela
                linhas_montagem.append("") ############ Atribuicao  #### olhar com a Rafaela
                linhas_montagem.append(f"ND {dados_linha['descricao']}") #Histórico
                
                 ############## Fim das Linhas
                sequencial += 1
                linhas_temp.append(linhas_montagem)
        self.dados_do_formulario_transferencia = linhas_temp
    
    def salvar_planilha(self):
        wb = openpyxl.load_workbook("MODELO BATCH INPUT.xlsx")
        ws = wb.active
        
        for x in range(10000):
            ws.delete_rows(2)
        
        for dados in self.dados_do_formulario_transferencia:
            ws.append(dados)

        options = {}
        options['defaultextension'] = ".xlsx"
        options['filetypes'] = [("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")]
        options['initialfile'] = "MODELO BATCH INPUT.xlsx"
        arquivo_salvar = filedialog.asksaveasfilename(**options)        
        wb.save(arquivo_salvar)



if __name__ == "__main__":
    configuracoes = Config()
    robo = Robo()
    robo.carregar_cadastro_de_empresas()
    robo.listar_arquivos()
    robo.carregar_arquivos_da_lista()
    robo.salvar_planilha()

    #print("############################################################")
    #divi_origem = robo.dados_do_formulario_transferencia[0]['divisao_origem']
    #print("############################################################")
    #print(robo.cadastro_de_empresas[robo.cadastro_de_empresas['Divisão'] == divi_origem]['Empresa'].values[0])
    